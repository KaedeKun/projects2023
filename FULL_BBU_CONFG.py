import json
from netaddr import IPNetwork
import pandas as pd
import ipaddress
import csv
import openpyxl
from colorama import Fore,Style
from jinja2 import Template
import time
import re
from tqdm import tqdm

wb = openpyxl.load_workbook('opt/EXP_980_Excel_BaseDatos_BBUs_Eth_FINAL.xlsx')

def cleaning():
	wb = openpyxl.load_workbook('opt/EXP_980_Excel_BaseDatos_BBUs_Eth_FINAL.xlsx')
	for wks in wb.worksheets:
		for row in wks["G2":"G321"]:
			for cell in row:
				cell.value = None
	
	wb.save('opt/EXP_980_Excel_BaseDatos_BBUs_Eth_FINAL.xlsx')

def add_description():
    num_bbu = []
    descr_bbu = []
    with open('templates/description_bbu_source.txt','r') as f:
        bbu_descr = f.readlines()
        wb.active = 0
        for ii in bbu_descr:
            separador = (ii.split("-"))
            num_bbu = separador[0]
            descr_bbu = separador[1].replace("\n","")
            for cell in wb.active['C']:
                if num_bbu in cell.value:
                    wb.active.cell(cell.row, cell.column+4).value="["+descr_bbu+"]"
                       
            wb.save('opt/EXP_980_Excel_BaseDatos_BBUs_Eth_FINAL.xlsx')
     
def suma3erocteto(segmento):
        return(segmento+1310720)

def formateoIP(direccion_bbu):
    return((json.dumps(direccion_bbu, indent=4, sort_keys=True, default=int))[1:-3])

print(Fore.LIGHTBLUE_EX+'TEMPLATE FULL BBUS (Eth-Trunks)')
print(Style.RESET_ALL)

while True:
    region_router = (input(Fore.LIGHTGREEN_EX+'Ingresa la región del Router: '+Style.RESET_ALL))
    if re.match(r"^\d+$", region_router):
            numero = int(region_router)
            if 1 <= numero <= 15:
                break
            else:
                print(Fore.RED+"ingrese región válida")
                continue
    else:
        print(Fore.RED+"Ingrese valor numerico valido")
        continue

while True:
    num_router  = (input(Fore.LIGHTGREEN_EX+'Ingresa el numero del Router: '+Style.RESET_ALL))
    if re.match(r"^\d+$", num_router):
        break
    else:
        print(Fore.RED+"Ingresa valores numéricos")
        continue

nombre_router = (input(Fore.LIGHTGREEN_EX+'Ingresa el nombre del Router: '+Style.RESET_ALL))
print(Style.RESET_ALL)
baseBBUDir = IPNetwork('11.'+region_router+'.'+num_router+'.0/24')

#PROCESO DE LIMPIEZA DE DESCRIPTION2
cleaning()
#-----------------------------------
while True:
    opcion = input(Fore.LIGHTGREEN_EX+'Desea agregar archivo de descripción? S/N: '+Style.RESET_ALL)
    if opcion == 'S' or opcion =='s':
        add_description()
        break
    elif opcion == 'N' or opcion =='n':
        print(Fore.RED+'No se agregan descripciones'+Style.RESET_ALL)
        break
    else:
        print(Fore.RED+'Valor no válido'+Style.RESET_ALL)
        continue

#--------------------------------------------------------------------------------------------
# SEGMENTACION Direccionamiento BBUs 980C ( O&M + Servicios )
BBUContador = 1
baseBBUDir = list(baseBBUDir.subnet(30))[0:64]
with open('opt/EXP_980_direcciones_raw_bbu.txt','w+',encoding='utf-8') as final_dirbbu:
    for ips in baseBBUDir:
        barra30bbu = list(ips.subnet(32))[1] #OBTIENE LA PRIMERA IP UTIL DEL SEGMENTO
        final_final = json.dumps(barra30bbu, indent=4, sort_keys=True, default=str)[1:-4] #FORMATO SIN CARACTERES EXTRAÑOS
        final_dirbbu.write(final_final+'\n') # GRABA LA  PRIMERA IP UTIL DE CADA SEGMENTO (8 PRIMEROS SEGMENTOS)
    BBUContador+=1
#--------------------------------------------------------------------------------------------
#Lectura de  lineas del archivo de IPs
# #-----------------------------------------------------------------------------
cont_oym = 2
cont_iub = 3
cont_s1 = 4
cont_s1_5g = 5 
with open('opt/EXP_980_direcciones_raw_bbu.txt','r') as f:
    lines = f.readlines()

for devices in tqdm(lines,desc="CONFIGURANDO TRONCALES",position=0, unit=" Eths",total=len(lines), colour="green"):
    BBU = devices
    OyM_BBU = formateoIP(BBU)
    IuB_BBU = suma3erocteto(ipaddress.ip_address(OyM_BBU))
    s1_BBU  = suma3erocteto(IuB_BBU)
    s1_5g_BBU = suma3erocteto(s1_BBU)

    with pd.ExcelWriter('opt/EXP_980_Excel_BaseDatos_BBUs_Eth_FINAL.xlsx', mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        pd.DataFrame([OyM_BBU]).to_excel     (writer, sheet_name='Excel_BaseDatos_BBUs', startcol=4, startrow=cont_oym, index=False, header=False)
        pd.DataFrame([IuB_BBU]).to_excel     (writer, sheet_name='Excel_BaseDatos_BBUs', startcol=4, startrow=cont_iub, index=False, header=False)
        pd.DataFrame([s1_BBU]).to_excel      (writer, sheet_name='Excel_BaseDatos_BBUs', startcol=4, startrow=cont_s1, index=False, header=False)
        pd.DataFrame([s1_5g_BBU]).to_excel   (writer, sheet_name='Excel_BaseDatos_BBUs', startcol=4, startrow=cont_s1_5g, index=False, header=False)
    cont_oym+=5
    cont_iub+=5
    cont_s1+=5
    cont_s1_5g+=5
    time.sleep(0.0000000001)

# #------------------------------------------------------------------------------
# #TRANSFORMACION DE ARCHIVO EXCEL A CSV  para ser manipulado por jinja2
mostrar = pd.DataFrame(pd.read_excel('opt/EXP_980_Excel_BaseDatos_BBUs_Eth_FINAL.xlsx'))
mostrar.to_csv('opt/EXP_980_BD_BBUs.csv', index=None, header=True, decimal='')
tranformacion_excel = pd.DataFrame(pd.read_csv('opt/EXP_980_BD_BBUs.csv'))

source_file = 'opt/EXP_980_BD_BBUs.csv'
BBU_Template_file ='templates/final_template_jinja_interfaces_910_bbu_eth_csv.j2'

with open (BBU_Template_file) as f:
        BBU_Template_file = Template(f.read(), keep_trailing_newline=True)

interface_configs = ""

with open(source_file) as s:
    reader = csv.DictReader(s)
    for row in reader:
        interface_config  = BBU_Template_file.render(
            name = row['name'],
            vlan = row['vlan'][:-1],
            description = row['description'],
            vpn_instance = row['vpn_instance'],
            ip_address = row['ip_address'],
            description2 = row['description2'],
        )
        interface_configs+= interface_config

with open(f'BBU_CONFIG_FINAL_{nombre_router}.txt','w') as lectura:
        lectura_final = lectura.write(interface_configs)

print(Fore.GREEN+Style.BRIGHT,f"LISTO!,", end="")
print(Fore.GREEN+Style.BRIGHT,f"CONFIGURACION PARA",Fore.LIGHTBLUE_EX,nombre_router,Fore.GREEN+Style.BRIGHT,"FULL BBU OK", end="")
print(Style.RESET_ALL)
# #--------------------------------------------------------------------------------------------------------
# ##-------------------------------------------------------------------------------------------------------